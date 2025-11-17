import os
import re
import sqlite3
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from datetime import datetime, time
from typing import Any, Dict, List, Optional, Tuple, Set
from difflib import SequenceMatcher
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell


# Paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
IN_DIR = os.path.join(BASE_DIR, "Read_report")
OUT_DIR = os.path.join(BASE_DIR, "out")
DB_PATH = os.path.join(OUT_DIR, "Operation_VTB.sqlite")


# Target table and columns
TABLE_NAME = "Operation_VTB"
CPT_TABLE = "Current_Portfolio"
COLUMNS = [
    "Portfolio",
    "Event",
    "Date",
    "Name",
    "Index",
    "Symbol",
    "Quantity",
    "Faceunit",
    "Price",
    "Currency",
    "Sumtransaction",
    "NKD",
    "FeeTax",
    "Note",
]

DuplicateKey = Tuple[Any, Any, Any, Any, Any, Any, Any, Any]

PORTFOLIO_MAP = {
    "124JAU": "124JAU STANDART",
    "1UJ8S": "1UJ8S KOPILKA",
    "124JAV": "124JAV IIS-3",
}


# Helpers
def ensure_dirs() -> None:
    os.makedirs(OUT_DIR, exist_ok=True)


def connect_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db(conn: sqlite3.Connection) -> None:
    # Note: SQLite accepts DATETIME as TEXT affinity; мы сохраняем ISO-строки.
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {TABLE_NAME} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            Portfolio TEXT NOT NULL,
            Event TEXT NOT NULL,
            Date DATETIME NOT NULL,
            Name TEXT,
            "Index" TEXT,
            Symbol TEXT,
            Quantity REAL,
            Faceunit TEXT,
            Price REAL,
            Currency TEXT,
            Sumtransaction REAL,
            NKD REAL,
            FeeTax REAL,
            Note TEXT
        )
        """
    )
    # Индекс для быстрого поиска дубликатов
    conn.execute(
        f"""
        CREATE INDEX IF NOT EXISTS idx_{TABLE_NAME}_dupe
        ON {TABLE_NAME} (Portfolio, Event, Date, Symbol, Quantity, Faceunit, Price, Sumtransaction)
        """
    )
    # Таблица текущего портфеля
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {CPT_TABLE} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            Portfolio TEXT NOT NULL,
            ReportDate DATE,
            SecType TEXT,
            Name TEXT,
            "Index" TEXT,
            Symbol TEXT,
            PlannedQty REAL,
            Faceunit TEXT,
            Price REAL,
            Nominal REAL,
            NKD REAL,
            CouponDate DATE,
            CouponRate REAL,
            PlannedValRub REAL
        )
        """
    )
    conn.commit()


def dec(value: Any) -> Optional[Decimal]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            d = Decimal(str(value))
        except InvalidOperation:
            return None
        return d.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
    s = str(value).strip().replace("\u00A0", "").replace(" ", "")
    if not s:
        return None
    try:
        d = Decimal(s)
        return d.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
    except InvalidOperation:
        return None


def dt(value: Any, force_date_only: bool = False) -> Optional[str]:
    # Возвращает ISO-строку YYYY-MM-DD HH:MM (без секунд).
    # Если только дата — ставим 23:00. Для force_date_only=True принудительно ставим 23:00,
    # даже если openpyxl вернул datetime с 00:00 (типично для колонок только с датой).
    if value is None:
        return None
    if isinstance(value, datetime):
        if force_date_only or (value.hour == 0 and value.minute == 0):
            value = datetime.combine(value.date(), time(23, 0))
        return value.strftime("%Y-%m-%d %H:%M")
    # openpyxl может дать date для excel-даты
    try:
        if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day") and not hasattr(value, "hour"):
            d = datetime.combine(value, time(23, 0))
            return d.strftime("%Y-%m-%d %H:%M")
    except Exception:
        pass
    s = str(value).strip()
    if not s:
        return None
    # Пробуем несколько форматов
    for fmt in [
        "%Y-%m-%d %H:%M",
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
    ]:
        try:
            t = datetime.strptime(s, fmt)
            if "%H:%M" in fmt:
                return t.strftime("%Y-%m-%d %H:%M")
            else:
                t = datetime.combine(t.date(), time(23, 0))
                return t.strftime("%Y-%m-%d %H:%M")
        except ValueError:
            continue
    return None


def cell_str(cell: Cell) -> str:
    v = cell.value
    if v is None:
        return ""
    return str(v).strip()


def norm_ccy(code: Optional[str]) -> str:
    s = (code or "").strip().upper()
    if s == "RUR":
        return "RUB"
    # Оставляем ISO, без лишних пробелов/кейса
    return s


def date_only_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().strftime("%Y-%m-%d")
    s = dt(value)
    if s:
        return s.split(" ")[0]
    return None

def find_portfolio(ws) -> Optional[str]:
    # Ищем строку с «№ субсчета:» и берём код справа в той же строке
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = cell_str(cell)
            if "№ субсчета:" in v:
                row_idx = cell.row
                for c in range(cell.column + 1, ws.max_column + 1):
                    code = cell_str(ws.cell(row=row_idx, column=c))
                    if code:
                        m = re.search(r"[A-Z0-9]+", code)
                        if m:
                            raw = m.group(0)
                            label = PORTFOLIO_MAP.get(raw, raw)
                            return label
                        return PORTFOLIO_MAP.get(code, code)
    return None


def find_section(ws, title: str) -> Optional[int]:
    # Возвращает номер строки, где ячейка содержит заголовок раздела (по вхождению, регистронезависимо)
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value:
                val = str(cell.value).strip()
                if title.lower() in val.lower():
                    return cell.row
    return None


def read_table_simple(ws, start_row: int, header: List[str], skip_after_header_rows: int = 0) -> List[List[Any]]:
    # Header ожидается на (start_row + 1)
    header_row = start_row + 1
    data_start = header_row + 1 + skip_after_header_rows
    rows: List[List[Any]] = []
    # Находим индексы колонок по строке заголовка
    cols: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        name = cell_str(ws.cell(row=header_row, column=c))
        if name in header:
            cols[name] = c
    # Должен быть хотя бы первый столбец
    if header[0] not in cols:
        return rows
    r = data_start
    while r <= ws.max_row:
        first_cell = cell_str(ws.cell(row=r, column=cols[header[0]]))
        # Стоп по полностью пустой строке (по всем известным колонкам)
        if all(cell_str(ws.cell(row=r, column=cols[h])) == "" for h in header if h in cols):
            break
        row_vals = []
        for col_name in header:
            cidx = cols.get(col_name)
            row_vals.append(ws.cell(row=r, column=cidx).value if cidx else None)
        rows.append(row_vals)
        r += 1
    return rows


def parse_opening_balances_v2(ws) -> Dict[str, Decimal]:
    title = "Сводная информация по субсчету Клиента"
    start = find_section(ws, title)
    if start is None:
        return {}
    header_row = start + 1
    # Робастный поиск колонок: учитываем возможные латинские буквы в заголовках
    cols: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        raw = cell_str(ws.cell(row=header_row, column=col)).strip().lower()
        if not raw:
            continue
        # Заменяем возможные латинские буквы на кириллицу для сравнения
        fixed = (
            raw.replace('c', 'с')
               .replace('o', 'о')
               .replace('p', 'р')
               .replace('e', 'е')
               .replace('x', 'х')
        )
        if 'опис' in fixed and 'desc' not in cols:
            cols['desc'] = col
        if (('сум' in fixed) or ('умм' in fixed)) and 'sum' not in cols:
            cols['sum'] = col
        if 'валют' in fixed and 'cur' not in cols:
            cols['cur'] = col
    if not cols.get('desc') or not cols.get('sum') or not cols.get('cur'):
        return {}
    balances: Dict[str, Decimal] = {}
    last_desc = ''
    r = header_row + 1
    while r <= ws.max_row:
        d = cell_str(ws.cell(row=r, column=cols['desc']))
        s = cell_str(ws.cell(row=r, column=cols['sum']))
        c = cell_str(ws.cell(row=r, column=cols['cur']))
        if d == '' and s == '' and c == '':
            break
        desc = d if d else last_desc
        last_desc = desc
        if desc.strip().lower() == 'входящий остаток денежных средств':
            amount = dec(ws.cell(row=r, column=cols['sum']).value)
            code = (c or '').upper()
            if code == 'RUR':
                code = 'RUB'
            if amount is not None and code:
                balances[code] = amount
        r += 1
    return balances






IGNORE_DDS_TYPES = {
    "Вознаграждение Брокера",
    "Сальдо расчётов по сделкам с ценными бумагами",
}


def map_event_dds(op_type: str, note: str, amount: Optional[Decimal]) -> Optional[str]:
    t = (op_type or "").strip()
    n = (note or "").strip()
    if t in IGNORE_DDS_TYPES:
        return None
    # Не импортируем агрегаты по валюте из ДДС
    if t == "Сальдо расчётов по сделкам с иностранной валютой":
        return None
    if t == "Зачисление денежных средств":
        return "CASH_IN"
    if t == "Списание денежных средств" and "Вывод денежных средств с брокерского счета" in n:
        return "CASH_OUT"
    if t in ("Купонный доход", "Дивиденды"):
        return "DIVIDEND"
    if t == "Погашение ценных бумаг":
        if "Част.погаш.номин." in n:
            return "AMORTISATION"
        if "Ден.ср-ва от погаш. номин.ст-ти" in n:
            return "REPAYMENT"
    if t == "НДФЛ":
        if amount is not None and amount < 0:
            return "FEE"
        if amount is not None and amount > 0:
            return "CASH_GAIN"
        return "FEE"
    # Прочее
    if amount is not None and amount < 0:
        return "CASH_EXPENSE"
    return t or None


def map_event_dds2(op_type: str, note: str, amount: Optional[Decimal]) -> Optional[str]:
    t = (op_type or "").strip()
    n = (note or "").strip()
    tl = t.lower()
    nl = n.lower()
    # Игнор: не импортируем эти строки из DDS
    if ("сальдо расч" in tl and ("ценн" in tl or "иностр" in tl)):
        return None
    if "вознаграждение брокера" in tl or "вознаграждение брокера" in nl:
        return None

    # Четкие правила маппинга по "Тип операции" и "Комментарий"
    if "зачисление денежных средств" in tl:
        return "CASH_IN"

    if "купонный доход" in tl:
        return "DIVIDEND"

    if "дивиденды" in tl:
        return "DIVIDEND"

    if "погашение ценных бумаг" in tl:
        # REPAYMENT: по сигнатурам в комментарии
        if ("ден.ср-ва от погаш. номин.ст-ти обл" in nl) or ("погашение номинальной стоимости облигаций" in nl):
            return "REPAYMENT"
        # AMORTISATION: частичное погашение
        if "част.погаш.номин. обл" in nl:
            return "AMORTISATION"
        # Если сигнатур нет — трактуем как прочий расход
        return "CASH_EXPENSE"

    if "списание денежных средств" in tl:
        return "CASH_OUT"

    if "ндфл" in tl:
        return "FEE"

    if "перераспределение доходов" in tl:
        return "CASH_EXPENSE"

    if "иные операции" in tl:
        if "част.погаш.номин. обл" in nl:
            return "DIVIDEND"
        return "CASH_EXPENSE"

    # По умолчанию — расход (неописанные правила)
    return "CASH_EXPENSE"


def parse_name_index_symbol(full: str) -> Tuple[str, str, str]:
    if not full:
        return "", "", ""
    parts = [p.strip() for p in str(full).split(",")]
    name = parts[0] if len(parts) > 0 else ""
    index = parts[1] if len(parts) > 1 else ""
    symbol = parts[2] if len(parts) > 2 else ""
    return name, index, symbol


def make_duplicate_key(row: Dict[str, Any]) -> DuplicateKey:
    return (
        row.get("Portfolio"),
        row.get("Event"),
        row.get("Date"),
        row.get("Symbol"),
        row.get("Quantity"),
        row.get("Faceunit"),
        row.get("Price"),
        row.get("Sumtransaction"),
    )


def fetch_existing_duplicate_keys(conn: sqlite3.Connection, portfolio: Optional[str]) -> Set[DuplicateKey]:
    """
    Returns the set of duplicate signatures that existed before the current import.
    Limiting by portfolio keeps the snapshot small and lets us allow duplicates inside a single file.
    """
    sql = f"""
        SELECT Portfolio, Event, Date, Symbol, Quantity, Faceunit, Price, Sumtransaction
        FROM {TABLE_NAME}
    """
    params: Tuple[Any, ...] = ()
    if portfolio:
        sql += " WHERE Portfolio=?"
        params = (portfolio,)
    cur = conn.execute(sql, params)
    return set(cur.fetchall())


def exists_in_db(conn: sqlite3.Connection, row: Dict[str, Any], baseline_only: Optional[Set[DuplicateKey]] = None) -> bool:
    key = make_duplicate_key(row)
    if baseline_only is not None:
        return key in baseline_only

    sql = f"""
        SELECT 1 FROM {TABLE_NAME}
        WHERE Portfolio=? AND Event=? AND Date=? AND Symbol=? AND Quantity IS ? AND Faceunit=? AND Price IS ? AND Sumtransaction IS ?
        LIMIT 1
    """
    cur = conn.execute(sql, key)
    return cur.fetchone() is not None


# Remains_* поля удалены: расчёт остатков в БД не ведётся


def update_remains(currents: Dict[str, Decimal], currency: Optional[str], event: str, sumtx: Optional[Decimal], feetax: Optional[Decimal], quantity: Optional[Decimal], name3: Optional[str], fx_side: Optional[str] = None) -> Dict[str, Decimal]:
    # currents: RUB, CNY, USD
    rr = currents.get("RUB", Decimal("0"))
    rcny = currents.get("CNY", Decimal("0"))
    rusd = currents.get("USD", Decimal("0"))
    sumtx = sumtx or Decimal("0")
    feetax = feetax or Decimal("0")
    quantity = quantity or Decimal("0")
    curcode = (currency or "").upper()
    if curcode == "RUR":
        curcode = "RUB"

    def add_cur(code: str, amount: Decimal):
        nonlocal rr, rcny, rusd
        if code == "RUB":
            rr = (rr + amount).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
        elif code == "CNY":
            rcny = (rcny + amount).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
        elif code == "USD":
            rusd = (rusd + amount).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)

    # Комиссии уменьшают баланс в валюте расчетов
    if feetax:
        add_cur(curcode, -feetax)

    if event in ("SELL", "CASH_IN", "DIVIDEND", "AMORTISATION", "REPAYMENT"):
        add_cur(curcode, sumtx)
    elif event == "BUY":
        add_cur(curcode, -sumtx)
    elif event == "CASH_OUT":
        add_cur(curcode, -sumtx)
    elif event == "FEE":
        # НДФЛ учитываем суммой операции (как в отчёте)
        add_cur(curcode, sumtx)
    elif event == "CASH_GAIN":
        add_cur(curcode, sumtx)
    elif event == "CASH_EXPENSE":
        add_cur(curcode, -sumtx)
    elif event == "CASH_CONVERT":
        # Конвертация: всегда работаем по стороне сделки, а не по знаку суммы
        tgt = (name3 or "").upper()
        side = (fx_side or "").strip()
        if curcode == "RUB" and tgt in ("CNY", "USD"):
            if side == "Покупка":
                # Покупка валюты → RUB уменьшается на сумму, иностранная растет на количество
                add_cur("RUB", -abs(sumtx))
                add_cur(tgt, abs(quantity))
            elif side == "Продажа":
                # Продажа валюты → RUB растет на сумму, иностранная уменьшается на количество
                add_cur("RUB", abs(sumtx))
                add_cur(tgt, -abs(quantity))
            else:
                # Нет стороны — меняем только RUB по сумме
                add_cur("RUB", sumtx)
        else:
            add_cur(curcode, sumtx)

    return {"RUB": rr, "CNY": rcny, "USD": rusd}


## Remains_* расчёты удалены\n

## Удалены вспомогательные аудит-функции по Remains_*

## Remains_* расчёты удалены\n

def build_instrument_index(conn: sqlite3.Connection, portfolio: str) -> Dict[str, Dict[str, Tuple[str, str, str]]]:
    idx_by = {"isin": {}, "index": {}, "name": {}}
    q = f"SELECT Name, \"Index\", Symbol FROM {TABLE_NAME} WHERE Portfolio=? AND (Event='BUY' OR Event='SELL')"
    for name, index, symbol in conn.execute(q, (portfolio,)):
        nm = (name or "").strip()
        ix = (index or "").strip()
        isin = (symbol or "").strip()
        triple = (nm, ix, isin)
        if isin:
            idx_by["isin"][isin.upper()] = triple
        if ix:
            idx_by["index"][ix.upper()] = triple
        if nm:
            key = re.sub(r"\s+", "", nm).upper()
            # Нормализуем имя для более устойчивого сравнения
            key_norm = normalize_name_key(nm)
            idx_by["name"][key] = triple
            idx_by["name"][key_norm] = triple
    return idx_by


def enrich_from_note(row: Dict[str, Any], note: str, idx_by: Dict[str, Dict[str, Tuple[str, str, str]]]) -> None:
    if row.get("Name") and row.get("Index") and row.get("Symbol"):
        return
    text = (note or "")
    up = text.upper()
    m_isin = re.search(r"RU[0-9A-Z]{9,10}", up)
    if m_isin:
        token = m_isin.group(0)
        triple = idx_by.get("isin", {}).get(token)
        if not triple:
            triple = idx_by.get("index", {}).get(token)
        if triple:
            row["Name"] = row.get("Name") or triple[0]
            row["Index"] = row.get("Index") or triple[1]
            row["Symbol"] = row.get("Symbol") or triple[2]
            return
        # Если по ISIN не нашли в индексе, заполним хотя бы Symbol и попробуем извлечь имя эмитента из текста
        row["Symbol"] = row.get("Symbol") or token
        # Попробуем вытащить имя между «облигациям/обл.» и ISIN/индексом
        m_name1 = re.search(r"облигац[ияем]{0,2}\s+([^,\n\r]+?)\s+ISIN", text, flags=re.IGNORECASE)
        if not m_name1:
            m_name1 = re.search(r"по\s+обл\.?\s+([^,\n\r]+?)\s+[A-Z0-9]{1,4}-", text, flags=re.IGNORECASE)
        if not m_name1:
            m_name1 = re.search(r"обл\.?\s+([^,\n\r]+?)\s+[A-Z0-9]{1,4}-", text, flags=re.IGNORECASE)
        if m_name1:
            issuer = m_name1.group(1).strip().strip('"“”«»')
            # попробуем найти по нормализованному имени среди BUY/SELL
            key_guess = normalize_name_key(issuer)
            triple2 = idx_by.get("name", {}).get(key_guess)
            if triple2:
                row["Name"] = row.get("Name") or triple2[0]
                row["Index"] = row.get("Index") or triple2[1]
                row["Symbol"] = row.get("Symbol") or triple2[2]
                return
            # если не нашли — хотя бы запишем Name эмитента
            row["Name"] = row.get("Name") or issuer
    m_idx = re.search(r"[A-Z0-9]{1,4}-[A-Z0-9]{2,}-[A-Z0-9-]{3,}", up)
    if m_idx:
        idx_token = m_idx.group(0)
        triple = idx_by.get("index", {}).get(idx_token)
        if triple:
            row["Name"] = row.get("Name") or triple[0]
            row["Index"] = row.get("Index") or triple[1]
            row["Symbol"] = row.get("Symbol") or triple[2]
            return
        else:
            # Если точного совпадения нет — хотя бы заполним Index значением из Note
            row["Index"] = row.get("Index") or idx_token
    # Доп. шаблон индекса без дефисов: например 26233RMFS
    m_idx2 = re.search(r"\b\d{3,}[A-Z]{2,}\b", up)
    if m_idx2:
        token = m_idx2.group(0)
        triple = idx_by.get("index", {}).get(token)
        if not triple:
            for key, tr in idx_by.get("index", {}).items():
                if token in key:
                    triple = tr
                    break
        if triple:
            row["Name"] = row.get("Name") or triple[0]
            row["Index"] = row.get("Index") or triple[1]
            row["Symbol"] = row.get("Symbol") or triple[2]
            return
        else:
            row["Index"] = row.get("Index") or token
    # Фаззи по имени: нормализуем текст заметки и ищем в индексе имён
    note_norm = normalize_name_key(up)
    best_key = None
    best_len = 0
    for key, triple in idx_by.get("name", {}).items():
        if not key or len(key) < 4:
            continue
        if key in note_norm or note_norm in key:
            lk = len(key)
            if lk > best_len:
                best_len = lk
                best_key = key
    if best_key:
        triple = idx_by["name"][best_key]
        row["Name"] = row.get("Name") or triple[0]
        row["Index"] = row.get("Index") or triple[1]
        row["Symbol"] = row.get("Symbol") or triple[2]
        return

    # Фаззи-поиск похожего имени: берём максимальную похожесть по нормализованным строкам
    best_key = None
    best_score = 0.0
    for key, triple in idx_by.get("name", {}).items():
        if not key or len(key) < 4:
            continue
        score = SequenceMatcher(None, note_norm, key).ratio()
        if score > best_score:
            best_score = score
            best_key = key
    if best_key and best_score >= 0.65:
        triple = idx_by["name"][best_key]
        row["Name"] = row.get("Name") or triple[0]
        row["Index"] = row.get("Index") or triple[1]
        row["Symbol"] = row.get("Symbol") or triple[2]
        return

    # Доп. fallback: перебор ключевых токенов из Note и сопоставление по подстроке
    tokens = re.findall(r"[А-ЯA-Z0-9]{3,}", up)
    token_best = None
    token_len_best = 0
    for t in tokens:
        tnorm = normalize_name_key(t)
        if len(tnorm) < 4:
            continue
        # Пробуем совпадение по индексам
        if tnorm in idx_by.get("index", {}):
            triple = idx_by["index"][tnorm]
            row["Name"] = row.get("Name") or triple[0]
            row["Index"] = row.get("Index") or triple[1]
            row["Symbol"] = row.get("Symbol") or triple[2]
            return
        for key in idx_by.get("name", {}).keys():
            if tnorm and (tnorm in key or key in tnorm):
                if len(tnorm) > token_len_best:
                    token_len_best = len(tnorm)
                    token_best = key
    if token_best:
        triple = idx_by["name"][token_best]
        row["Name"] = row.get("Name") or triple[0]
        row["Index"] = row.get("Index") or triple[1]
        row["Symbol"] = row.get("Symbol") or triple[2]
        return

def normalize_name_key(s: str) -> str:
    # Убираем юридические формы и служебные слова, дефисы/кавычки/точки/пробелы
    if not s:
        return ""
    s = s.upper()
    s = s.replace("Ё", "Е")
    # Удаляем юридические формы
    s = re.sub(r"\b(ПАО|ОАО|ЗАО|АО|ООО)\b", "", s)
    # Удаляем служебные слова часто встречающиеся в Note
    s = re.sub(r"\b(ПО|ЦЕННЫМ|ЦЕННЫЕ|БУМАГАМ|БУМАГИ|ГОД|НДС|НЕ|ОБЛ|УДЕРЖАН|НАЛОГ|В|РАЗМЕРЕ|РУБ)\b", "", s)
    # Убираем суффикс AO/АО
    s = s.replace("-AO", "").replace("-АО", "").replace("АО-", "").replace("AO-", "")
    # Удаляем всё кроме букв и цифр
    s = re.sub(r"[^0-9A-ZА-Я]", "", s)
    return s


def insert_row(conn: sqlite3.Connection, row: Dict[str, Any]) -> int:
    placeholders = ",".join(["?"] * len(COLUMNS))
    col_list = ",".join([f'"{c}"' if c == "Index" else c for c in COLUMNS])
    sql = f"INSERT INTO {TABLE_NAME} ({col_list}) VALUES ({placeholders})"
    # Значение Currency по умолчанию — RUR, если пусто
    if not (row.get("Currency") or "").strip():
        row["Currency"] = "RUR"
    values = [row.get(col) for col in COLUMNS]
    cur = conn.execute(sql, values)
    return cur.lastrowid


def _dump_row(ws, r: int, limit: int = 30) -> str:
    vals = []
    maxc = min(ws.max_column, limit)
    for c in range(1, maxc + 1):
        vals.append(cell_str(ws.cell(row=r, column=c)))
    return " | ".join(vals)


def debug_section(ws, title: str) -> None:
    s = find_section(ws, title)
    print(f"Debug: section '{title}' start_row={s}")
    if s:
        hdr = s + 1
        print(f"Debug: header row {hdr}: {_dump_row(ws, hdr)}")
        print(f"Debug: first data row {hdr+1}: {_dump_row(ws, hdr+1)}")


def process_dds(ws, portfolio: str) -> List[Dict[str, Any]]:
    title = "Движение денежных средств"
    start = find_section(ws, title)
    if start is None:
        return []
    header = ["Дата", "Сумма", "Валюта", "Тип операции", "Комментарий"]
    # После заголовка идёт строка с рынком → пропускаем 1 строку
    rows = read_table_simple(ws, start, header, skip_after_header_rows=1)
    results: List[Dict[str, Any]] = []
    for vals in rows:
        date_s = dt(vals[0], force_date_only=True)
        amount = dec(vals[1])
        currency = str(vals[2]).strip() if vals[2] is not None else ""
        op_type = str(vals[3]).strip() if vals[3] is not None else ""
        note = str(vals[4]).strip() if vals[4] is not None else ""
        # Более устойчивое сопоставление типов операций по подстрокам
        tl = op_type.lower()
        nl = note.lower()
        if (("комиссия" in tl or "комиссия" in nl) and ("валютн" in tl or "валютн" in nl) and ("заключени" in tl or "заключени" in nl or "проведен" in tl or "проведен" in nl)):
            continue
        event = map_event_dds2(op_type, note, amount)
        if event is None:
            continue  # пропускаем «Вознаграждение Брокера»

        # По правилу: если комментарий пуст, Name = Валюта
        name = currency if note == "" else ""
        index = ""
        symbol = ""
        nkd = None
        feetax = None

        if event in ("AMORTISATION", "REPAYMENT"):
            m_index = re.search(r"([A-Z0-9]{2,}-[A-Z0-9-]{4,})", note)
            if m_index:
                index = m_index.group(1)

        # Политика знаков суммы:
        # - FEE: записываем исходную сумму в Sumtransaction (знак как в отчёте), FeeTax не заполняем
        # - CASH_OUT и BUY: Sumtransaction отрицательная (по модулю исходной)
        # - Остальные события: Sumtransaction положительная (по модулю исходной)
        sum_for_row: Optional[float]
        fee_for_row: Optional[float]
        if event == "FEE":
            fee_for_row = None
            sum_for_row = float(amount) if amount is not None else None
        else:
            fee_for_row = None
            if amount is None:
                sum_for_row = None
            else:
                a = float(abs(amount))
        # Отмена принудительного изменения знака: берем как в отчете
        sum_for_row = float(amount) if amount is not None else None

        # Для FEE, CASH_OUT, CASH_IN заполняем Name/Index/Symbol значением валюты
        if event in ("FEE", "CASH_OUT", "CASH_IN"):
            name = currency
            index = currency
            symbol = currency

        row: Dict[str, Any] = {
            "Portfolio": portfolio,
            "Event": event,
            "Date": date_s,
            "Name": name,
            "Index": index,
            "Symbol": symbol,
            # Для денежных движений Quantity ставим 1
            "Quantity": float(dec(1) or Decimal("1")),
            "Faceunit": currency,
            "Price": None,
            "Currency": currency,
            "Sumtransaction": sum_for_row,
            "NKD": float(nkd) if nkd is not None else None,
            "FeeTax": float(feetax) if feetax is not None else fee_for_row,
            "Note": note,
        }
        results.append(row)
    return results


def _map_columns_by_contains(ws, header_row: int, patterns: Dict[str, List[str]]) -> Dict[str, int]:
    cols: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        raw = cell_str(ws.cell(row=header_row, column=c))
        low = raw.lower()
        for key, pats in patterns.items():
            if key in cols:
                continue
            for p in pats:
                if p in low:
                    cols[key] = c
                    break
    return cols


def find_report_date(ws) -> Optional[str]:
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = cell_str(cell)
            if "Дата формирования отчета" in v:
                row_idx = cell.row
                for c in range(cell.column + 1, ws.max_column + 1):
                    val = ws.cell(row=row_idx, column=c).value
                    if val is not None and str(val).strip() != "":
                        return date_only_str(val)
                return None
    return None


def process_current_portfolio(ws, portfolio: str) -> List[Dict[str, Any]]:
    title = "Отчёт об остатках ценных бумаг"
    start = find_section(ws, title)
    if start is None:
        return []
    header_row = start + 1
    # Заголовки могут быть многострочными — работаем по подстрокам
    def low_txt(r, c):
        return cell_str(ws.cell(row=r, column=c)).lower().replace("\n", " ")

    patterns = {
        "tris": ["наименование ценной бумаги", "isin"],
        "planned_qty": ["плановый исходящий остаток"],
        "faceunit": ["валюта", "номинала"],
        "price": ["цена", "%"],
        "nominal": [" номинал"],
        "nkd": ["нкд", "на конец периода"],
        "coupon_date": ["дата выплаты", "дата погашения"],
        "coupon_rate": ["ставка купона"],
        "planned_val_rub": ["оценка планового исходящего остатка", "руб"],
    }
    cols = _map_columns_by_contains(ws, header_row, patterns)
    # Уточнение колонки Nominal: избегаем совпадения с фразой "в валюте номинала" / "НКД"
    hdr_low = {c: cell_str(ws.cell(row=header_row, column=c)).lower().replace("\n"," ") for c in range(1, ws.max_column+1)}
    def is_nominal_header(txt: str) -> bool:
        return ("номинал" in txt) and ("в валюте" not in txt) and ("валют" not in txt) and ("нкд" not in txt)
    if cols.get("nominal"):
        h = hdr_low.get(cols["nominal"], "")
        if not is_nominal_header(h):
            # попробуем найти более подходящую колонку
            for c, low in hdr_low.items():
                if is_nominal_header(low):
                    cols["nominal"] = c
                    break
    else:
        for c, low in hdr_low.items():
            if is_nominal_header(low):
                cols["nominal"] = c
                break

    # Отладочный вывод маппинга колонок текущего портфеля
    try:
        print(f"Debug: CPT cols mapped: {cols}")
    except Exception:
        pass
    # Дата формирования отчета
    rep_date = find_report_date(ws)

    results: List[Dict[str, Any]] = []
    sec_type: Optional[str] = None
    sec_type_tokens = {"еврооблигация", "облигация", "пай", "акция"}
    r = header_row + 1
    while r <= ws.max_row:
        # Текстовые значения в ключевых местах
        first_val = cell_str(ws.cell(row=r, column=1)).strip()
        tris_val = cell_str(ws.cell(row=r, column=cols.get("tris", 1))).strip() if cols.get("tris") else ""

        # Окончание таблицы: строка с "ИТОГО:" в первом столбце или в столбце tris
        if (first_val.upper() == "ИТОГО:" or tris_val.upper() == "ИТОГО:"):
            break

        # Определение блоков типа ЦБ: значение SecType находится в колонке tris,
        # и это одна из: ЕВРООБЛИГАЦИЯ/ОБЛИГАЦИЯ/ПАЙ/АКЦИЯ, при этом прочие данные в строке отсутствуют
        tris_low = tris_val.lower()
        if tris_low in sec_type_tokens:
            # проверим, что остальные ключевые колонки пустые (чтобы не принять строку-данные за заголовок типа)
            others_empty = True
            for key in ("planned_qty","faceunit","price","nominal","nkd","coupon_date","coupon_rate","planned_val_rub"):
                c = cols.get(key)
                if c and cell_str(ws.cell(row=r, column=c)) != "":
                    others_empty = False
                    break
            if others_empty:
                sec_type = tris_val
                r += 1
                continue

        # Признак возможного окончания блока — полностью пустые ключевые колонки; проверим следующую строку
        if cols.get("tris") and tris_val == "" and first_val == "":
            nxt = r + 1
            if nxt > ws.max_row or (cols.get("tris") and cell_str(ws.cell(row=nxt, column=cols["tris"])) == ""):
                break
        # Читаем строку данных
        tris_raw = cell_str(ws.cell(row=r, column=cols.get("tris", 0))) if cols.get("tris") else ""
        if not tris_raw.strip():
            r += 1
            continue
        name, index, symbol = parse_name_index_symbol(tris_raw)
        planned_qty = dec(ws.cell(row=r, column=cols.get("planned_qty", 0)).value) if cols.get("planned_qty") else None
        faceunit = cell_str(ws.cell(row=r, column=cols.get("faceunit", 0))) if cols.get("faceunit") else ""
        price = dec(ws.cell(row=r, column=cols.get("price", 0)).value) if cols.get("price") else None
        nominal = dec(ws.cell(row=r, column=cols.get("nominal", 0)).value) if cols.get("nominal") else None
        nkd = dec(ws.cell(row=r, column=cols.get("nkd", 0)).value) if cols.get("nkd") else None
        coupon_date = ws.cell(row=r, column=cols.get("coupon_date", 0)).value if cols.get("coupon_date") else None
        coupon_rate = dec(ws.cell(row=r, column=cols.get("coupon_rate", 0)).value) if cols.get("coupon_rate") else None
        planned_val_rub = dec(ws.cell(row=r, column=cols.get("planned_val_rub", 0)).value) if cols.get("planned_val_rub") else None

        # Пропуск строк с плановым исходящим остатком = 0
        if planned_qty is not None and planned_qty == Decimal("0"):
            r += 1
            continue

        row_out = {
            "Portfolio": portfolio,
            "ReportDate": rep_date,
            "SecType": sec_type,
            "Name": name,
            "Index": index,
            "Symbol": symbol,
            "PlannedQty": float(planned_qty) if planned_qty is not None else None,
            "Faceunit": faceunit,
            "Price": float(price) if price is not None else None,
            "Nominal": float(nominal) if nominal is not None else None,
            "NKD": float(nkd) if nkd is not None else None,
            "CouponDate": date_only_str(coupon_date),
            "CouponRate": float((coupon_rate or Decimal("0"))) if coupon_rate is not None else None,
            "PlannedValRub": float(planned_val_rub) if planned_val_rub is not None else None,
        }
        results.append(row_out)
        r += 1
    return results


def upsert_current_portfolio(conn: sqlite3.Connection, ws, portfolio: str) -> None:
    rows = process_current_portfolio(ws, portfolio)
    if not rows:
        return
    conn.execute(f"DELETE FROM {CPT_TABLE} WHERE Portfolio=?", (portfolio,))
    cols = [
        "Portfolio","ReportDate","SecType","Name","Index","Symbol",
        "PlannedQty","Faceunit","Price","Nominal","NKD","CouponDate","CouponRate","PlannedValRub",
    ]
    placeholders = ",".join(["?"]*len(cols))
    sql = f"INSERT INTO {CPT_TABLE} (" + ",".join([f'"{c}"' if c=="Index" else c for c in cols]) + ") VALUES ("+placeholders+")"
    for r in rows:
        conn.execute(sql, [r.get(c) for c in cols])
    conn.commit()




def process_securities(ws, portfolio: str) -> List[Dict[str, Any]]:
    title = "Заключенные в отчетном периоде сделки с ценными бумагами"
    start = find_section(ws, title)
    if start is None:
        return []
    header_row = start + 1
    patterns = {
        "name": ["наименование ценной", "isin"],
        "dt": ["дата и время"],
        "side": ["вид сделки"],
        "qty": ["количество", "шт"],
        "face": ["валюта цены", "номинала"],
        "price": ["цена"],
        "cur": ["валюта расчетов"],
        "sum": ["сумма сделки", "валюте расчетов"],
        "nkd": ["нкд по сделке"],
        "fee_settle": ["комиссия банка за расчет"],
        "fee_trade": ["комиссия банка за заключение"],
        "comment": ["комментарий"],
    }
    cols = _map_columns_by_contains(ws, header_row, patterns)
    # Доп. корректировка: если NKD совпал с колонкой суммы, пытаемся найти заголовок, содержащий именно "нкд по"
    if cols.get("nkd") and cols.get("sum") and cols["nkd"] == cols["sum"]:
        for c in range(1, ws.max_column + 1):
            raw1 = cell_str(ws.cell(row=header_row, column=c)).lower().replace("`n", " ")
            raw2 = cell_str(ws.cell(row=header_row + 1, column=c)).lower().replace("`n", " ") if header_row + 1 <= ws.max_row else ""
            if ("нкд по" in raw1) or ("нкд по" in raw2):
                cols["nkd"] = c
                break
    # Fallback: если колонка NKD не найдена по шаблонам, пробуем искать по подстроке 'нкд'
    if not cols.get("nkd"):
        for c in range(1, ws.max_column + 1):
            raw = cell_str(ws.cell(row=header_row, column=c)).lower()
            if "нкд" in raw:
                cols["nkd"] = c
                break
    print(f"Debug: SEC cols mapped: {cols}")
    if not cols.get("name") or not cols.get("dt"):
        return []
    r = header_row + 1
    results: List[Dict[str, Any]] = []
    while r <= ws.max_row:
        # стоп по полностью пустой строке по ключевым колонкам
        if all(cell_str(ws.cell(row=r, column=cols[k])) == "" for k in ("name", "dt") if k in cols):
            break
        name_raw = cell_str(ws.cell(row=r, column=cols["name"])) if cols.get("name") else ""
        date_s = dt(ws.cell(row=r, column=cols["dt"]).value)
        side = cell_str(ws.cell(row=r, column=cols.get("side", 0))) if cols.get("side") else ""
        qty = dec(ws.cell(row=r, column=cols.get("qty", 0)).value) if cols.get("qty") else None
        faceunit = cell_str(ws.cell(row=r, column=cols.get("face", 0))) if cols.get("face") else ""
        price = dec(ws.cell(row=r, column=cols.get("price", 0)).value) if cols.get("price") else None
        currency = cell_str(ws.cell(row=r, column=cols.get("cur", 0))) if cols.get("cur") else ""
        sum_col = cols.get("sum")
        sumtx = dec(ws.cell(row=r, column=sum_col).value) if sum_col else None
        nkd = dec(ws.cell(row=r, column=cols.get("nkd", 0)).value) if cols.get("nkd") else None
        # Если колонка NKD не определена корректно, попробуем взять первое числовое значение справа от суммы
        if sum_col and (
            (cols.get("nkd") and cols["nkd"] == sum_col) or nkd is None
        ):
            for c in range(sum_col + 1, ws.max_column + 1):
                v = dec(ws.cell(row=r, column=c).value)
                if v is not None:
                    nkd = v
                    break
        fee1 = dec(ws.cell(row=r, column=cols.get("fee_settle", 0)).value) if cols.get("fee_settle") else None
        fee2 = dec(ws.cell(row=r, column=cols.get("fee_trade", 0)).value) if cols.get("fee_trade") else None
        note = cell_str(ws.cell(row=r, column=cols.get("comment", 0))) if cols.get("comment") else ""

        name, index, symbol = parse_name_index_symbol(name_raw)
        event = "BUY" if side == "Покупка" else ("SELL" if side == "Продажа" else side)
        feetax = (fee1 or Decimal("0")) + (fee2 or Decimal("0"))

        # Знак суммы: BUY отрицательная, SELL положительная
        sum_for_row = float(sumtx) if sumtx is not None else None

        row: Dict[str, Any] = {
            "Portfolio": portfolio,
            "Event": event,
            "Date": date_s,
            "Name": name,
            "Index": index,
            "Symbol": symbol,
            "Quantity": float(qty) if qty is not None else None,
            "Faceunit": faceunit,
            "Price": float(price) if price is not None else None,
            "Currency": currency,
            "Sumtransaction": sum_for_row,
            "NKD": float(nkd) if nkd is not None else None,
            "FeeTax": float(feetax) if feetax is not None else None,
            "Note": note,
        }
        results.append(row)
        r += 1
    return results


def process_fx(ws, portfolio: str) -> List[Dict[str, Any]]:
    title = "Заключенные в отчетном периоде сделки с иностранной валютой"
    start = find_section(ws, title)
    if start is None:
        return []
    header_row = start + 1
    patterns = {
        "instr": ["финансовый инструмент"],
        "dt": ["дата и время"],
        "side": ["вид сделки"],
        "qty": ["количество", "шт"],
        "price": ["цена"],
        "cur": ["валюта расчетов"],
        "sum": ["сумма сделки", "валюте расчетов"],
        "fee_settle": ["комиссия банка за расчет"],
        "fee_trade": ["комиссия банка за заключение"],
        "comment": ["комментарий"],
    }
    cols = _map_columns_by_contains(ws, header_row, patterns)
    print(f"Debug: FX cols mapped: {cols}")
    if not cols.get("instr") or not cols.get("dt"):
        return []
    r = header_row + 1
    results: List[Dict[str, Any]] = []
    while r <= ws.max_row:
        if all(cell_str(ws.cell(row=r, column=cols[k])) == "" for k in ("instr", "dt") if k in cols):
            break
        instr = cell_str(ws.cell(row=r, column=cols.get("instr", 0))) if cols.get("instr") else ""
        first3 = instr[:3] if instr else ""
        second3 = instr[3:6] if instr and len(instr) >= 6 else (instr[-3:] if instr else "")
        date_s = dt(ws.cell(row=r, column=cols["dt"]).value)
        qty = dec(ws.cell(row=r, column=cols.get("qty", 0)).value) if cols.get("qty") else None
        price = dec(ws.cell(row=r, column=cols.get("price", 0)).value) if cols.get("price") else None
        currency = cell_str(ws.cell(row=r, column=cols.get("cur", 0))) if cols.get("cur") else ""
        sum_col = cols.get("sum")
        sumtx = dec(ws.cell(row=r, column=sum_col).value) if sum_col else None
        fee1 = dec(ws.cell(row=r, column=cols.get("fee_settle", 0)).value) if cols.get("fee_settle") else None
        fee2 = dec(ws.cell(row=r, column=cols.get("fee_trade", 0)).value) if cols.get("fee_trade") else None
        note = cell_str(ws.cell(row=r, column=cols.get("comment", 0))) if cols.get("comment") else ""
        side = cell_str(ws.cell(row=r, column=cols.get("side", 0))) if cols.get("side") else ""
        feetax = (fee1 or Decimal("0")) + (fee2 or Decimal("0"))
        # Правило знаков по требованию:
        # Покупка: Name=первые 3 буквы, Quantity>0, Sumtransaction<0
        # Продажа: Name=вторые 3 буквы, Quantity<0, Sumtransaction>0
        name_val = (first3 if side == "���㯪�" else (second3 if side == "�த���" else first3))
        qty_val = float(qty) if qty is not None else None
        sum_for_row = float(sumtx) if sumtx is not None else None
        if side == "Покупка":
            name_val = first3
            
        elif side == "Продажа":
            name_val = second3
            

        row: Dict[str, Any] = {
            "Portfolio": portfolio,
            "Event": "CASH_CONVERT",
            "Date": date_s,
            "Name": name_val,
            "Index": name_val,
            "Symbol": name_val,
            "Quantity": qty_val,
            # Faceunit всегда = первые 3 символа (базовая валюта количества)
            "Faceunit": first3,
            "Price": float(price) if price is not None else None,
            "Currency": currency,
            "Sumtransaction": sum_for_row,
            "NKD": None,
            "FeeTax": float(feetax) if feetax is not None else None,
            "Note": note,
        }
        # Сохраним сторону сделки во вспомогательном поле для update_remains через Note (префикс)
        if side:
            row["Note"] = f"{side}|{note}" if note else side
        results.append(row)
        r += 1
    return results


# Opening balance validation can be disabled by flag below
OPENING_BALANCE_CHECK_ENABLED = False

def validate_opening_balances(conn: sqlite3.Connection, ws, portfolio: str) -> None:
    if not OPENING_BALANCE_CHECK_ENABLED:
        return
    reported = parse_opening_balances_v2(ws)
    has_rows = conn.execute(
        f"SELECT 1 FROM {TABLE_NAME} WHERE Portfolio=? LIMIT 1", (portfolio,)
    ).fetchone() is not None
    if not has_rows:
        for code in ("RUB", "CNY", "USD"):
            val = reported.get(code, Decimal("0"))
            if val != Decimal("0"):
                raise RuntimeError("Входящий остатот денежных средств не совпадает.")
        return
    expected = {"RUB": Decimal("0"), "CNY": Decimal("0"), "USD": Decimal("0")}
    for code in ("RUB", "CNY", "USD"):
        rep = reported.get(code)
        exp = expected.get(code, Decimal("0"))
        if rep is None:
            if exp != Decimal("0"):
                raise RuntimeError("Входящий остатот денежных средств не совпадает.")
        elif rep != exp:
            raise RuntimeError("Входящий остатот денежных средств не совпадает.")


def move_processed(src_path: str) -> None:
    fname = os.path.basename(src_path)
    dst = os.path.join(BASE_DIR, fname)
    if os.path.exists(dst):
        dst = os.path.join(BASE_DIR, f"D_{fname}")
    os.replace(src_path, dst)


def process_file(path: str, conn: sqlite3.Connection) -> Tuple[int, int]:
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    portfolio = find_portfolio(ws) or "UNKNOWN"
    # Snapshot duplicate keys for rows that existed before this file import.
    existing_dupe_keys = fetch_existing_duplicate_keys(conn, portfolio)

    # Opening balances check (disabled via flag by default)
    validate_opening_balances(conn, ws, portfolio)

    # Обновим таблицу текущего портфеля для данного портфеля
    try:
        upsert_current_portfolio(conn, ws, portfolio)
    except Exception as e:
        print(f"Warning: не удалось обновить текущий портфель для {portfolio}: {e}")

    # Parse sections
    rows_dds = process_dds(ws, portfolio)
    rows_sec = process_securities(ws, portfolio)
    rows_fx = process_fx(ws, portfolio)
    rows_all: List[Dict[str, Any]] = rows_dds + rows_sec + rows_fx

    # Enrich identifiers using BUY/SELL index
    idx_by = build_instrument_index(conn, portfolio)
    for rs in rows_sec:
        if rs.get("Event") in ("BUY", "SELL"):
            nm = (rs.get("Name") or "").strip()
            ix = (rs.get("Index") or "").strip()
            isin = (rs.get("Symbol") or "").strip()
            triple = (nm, ix, isin)
            if isin:
                idx_by["isin"][isin.upper()] = triple
            if ix:
                idx_by["index"][ix.upper()] = triple
            if nm:
                key = re.sub(r"\s+", "", nm).upper()
                idx_by["name"][key] = triple
                idx_by["name"][normalize_name_key(nm)] = triple

    for r in rows_all:
        if r.get("Event") in ("AMORTISATION", "REPAYMENT", "DIVIDEND"):
            if not (r.get("Name") and r.get("Index") and r.get("Symbol")):
                enrich_from_note(r, r.get("Note") or "", idx_by)

    # Sort by date
    rows_all.sort(key=lambda r: (r.get("Date") or ""))

    inserted: List[Dict[str, Any]] = []
    dup_skipped = 0
    for row in rows_all:
        for k in ("Quantity", "Price", "Sumtransaction", "NKD", "FeeTax"):
            if row.get(k) is None:
                continue
            val = dec(row[k])
            if val is None:
                continue
            row[k] = float(val)
        if exists_in_db(conn, row, baseline_only=existing_dupe_keys):
            dup_skipped += 1
            continue
        insert_row(conn, row)
        inserted.append(row)

    conn.commit()
    return len(rows_all), len(inserted)


def main() -> None:
    ensure_dirs()
    conn = connect_db()
    init_db(conn)

    files = [
        os.path.join(IN_DIR, f)
        for f in os.listdir(IN_DIR)
        if f.lower().endswith('.xlsx')
    ] if os.path.isdir(IN_DIR) else []

    total_found = 0
    total_inserted = 0
    for fpath in files:
        try:
            found, inserted = process_file(fpath, conn)
            total_found += found
            total_inserted += inserted
            move_processed(fpath)
            print(f"Processed {os.path.basename(fpath)}: found={found}, inserted={inserted}")
        except Exception as e:
            print(f"Error processing {os.path.basename(fpath)}: {e}")

    print(f"Done. Total found={total_found}, inserted={total_inserted}")


if __name__ == '__main__':
    main()





