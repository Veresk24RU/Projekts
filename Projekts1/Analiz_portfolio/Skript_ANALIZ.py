import argparse
import os
import sqlite3
import shutil
import tempfile
from typing import Dict, List, Sequence, Tuple, Optional, Union
from datetime import datetime, date, time
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP


def _quote_ident(name: str) -> str:
    return '"' + str(name).replace('"', '""') + '"'


def _norm_table_name(name: str) -> str:
    return str(name).strip()


def _try_import_pandas():
    try:
        import pandas as pd  # type: ignore
        return pd
    except Exception:
        return None


def _try_import_openpyxl():
    try:
        import openpyxl  # type: ignore
        return openpyxl
    except Exception:
        return None


def _infer_sqlite_type_from_pd_dtype(pd_dtype) -> str:
    kind = getattr(pd_dtype, "kind", None)
    if kind in ("i", "u"):
        return "INTEGER"
    if kind == "f":
        return "REAL"
    if kind == "b":
        return "INTEGER"
    if kind in ("M", "m"):
        return "TEXT"
    return "TEXT"


def read_schema_from_excel(schema_path: str, sheet: Optional[Union[str, int]] = None) -> Tuple[List[str], Dict[str, str]]:
    """
    Возвращает (список_столбцов, типы_по_столбцам). Типы — SQLite.
    Если доступен pandas — типы инферятся из данных; иначе все TEXT.
    """
    pd = _try_import_pandas()
    if pd is not None:
        sheet_arg = 0 if sheet is None else sheet
        df = pd.read_excel(schema_path, sheet_name=sheet_arg, nrows=1000)
        cols = [str(c) for c in list(df.columns)]
        types = {str(col): _infer_sqlite_type_from_pd_dtype(dtype) for col, dtype in df.dtypes.items()}
        return cols, types

    ox = _try_import_openpyxl()
    if ox is None:
        raise RuntimeError("Не удалось импортировать pandas или openpyxl. Установите пакет: pip install pandas openpyxl")
    wb = ox.load_workbook(schema_path, read_only=True, data_only=True)
    ws = wb[sheet] if isinstance(sheet, str) else (wb.worksheets[sheet] if isinstance(sheet, int) else wb.active)
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    cols = [str(c) if c is not None else f"col_{i+1}" for i, c in enumerate(header)]
    types = {c: "TEXT" for c in cols}
    wb.close()
    return cols, types


def ensure_table(conn: sqlite3.Connection, table: str, columns: Sequence[str], types: Dict[str, str], if_exists: str) -> None:
    cur = conn.cursor()
    table_q = _quote_ident(_norm_table_name(table))
    if if_exists == "replace":
        cur.execute(f"DROP TABLE IF EXISTS {table_q}")
    cols_sql = ", ".join(f"{_quote_ident(c)} {types.get(c, 'TEXT')}" for c in columns)
    cur.execute(f"CREATE TABLE IF NOT EXISTS {table_q} ({cols_sql})")
    conn.commit()


def _short_hash(text: str) -> str:
    import hashlib
    return hashlib.sha1(text.encode("utf-8", errors="ignore")).hexdigest()[:10]


def ensure_unique_index(conn: sqlite3.Connection, table: str, unique_cols: Sequence[str] | None) -> Optional[str]:
    """Создаёт уникальный индекс по заданным столбцам. Возвращает имя индекса или None, если не требуется."""
    if not unique_cols:
        return None
    tbl_name = _norm_table_name(table)
    cols_tuple = tuple(unique_cols)
    cols_sig = ",".join(cols_tuple)
    idx_name = f"ux_{tbl_name}_{_short_hash(cols_sig)}"
    idx_q = _quote_ident(idx_name)
    table_q = _quote_ident(tbl_name)
    cols_q = ", ".join(_quote_ident(c) for c in cols_tuple)
    cur = conn.cursor()
    cur.execute(f"CREATE UNIQUE INDEX IF NOT EXISTS {idx_q} ON {table_q} ({cols_q})")
    conn.commit()
    return idx_name


def table_exists(conn: sqlite3.Connection, table: str) -> bool:
    cur = conn.cursor()
    cur.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name = ?",
        (_norm_table_name(table),),
    )
    return cur.fetchone() is not None


def get_db_table_schema(conn: sqlite3.Connection, table: str) -> Tuple[List[str], Dict[str, str]]:
    """Возвращает (список столбцов, типы) текущей таблицы в БД."""
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info({_quote_ident(_norm_table_name(table))})")
    cols: List[str] = []
    types: Dict[str, str] = {}
    for cid, name, ctype, notnull, dflt_value, pk in cur.fetchall():
        cols.append(str(name))
        types[str(name)] = str(ctype or "TEXT")
    return cols, types


def _to_float(val: object) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _round_decimal(val: Optional[float], digits: int = 6) -> Optional[float]:
    if val is None:
        return None
    try:
        d = Decimal(str(val)).quantize(Decimal(10) ** -digits, rounding=ROUND_HALF_UP)
        return float(d)
    except (InvalidOperation, ValueError):
        return val


def _normalize_datetime_str(val: object) -> Optional[str]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(val, date):
        return datetime.combine(val, time.min).strftime("%Y-%m-%d %H:%M:%S")
    s = str(val).strip()
    # Попытка ISO-парсинга
    try:
        # Пробуем разобрать как ISO, затем привести к формату с секундами
        dt = datetime.fromisoformat(s.replace("T", " "))
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        # fallback: если строка формата YYYY-MM-DD, добавим время 00:00:00
        if len(s) == 10 and s[4] == '-' and s[7] == '-':
            return f"{s} 00:00:00"
        return s


def _normalize_row_values(row: Dict[str, object], columns: Sequence[str]):
    out: List[object] = []
    for c in columns:
        v = row.get(c)
        cl = c.lower()
        # Нормализация ключевых полей для дедупликации
        if cl == "event":
            v = _norm_str(v).upper()
        elif cl == "symbol":
            v = _norm_str(v)
        elif cl == "date":
            v = _normalize_datetime_str(v)
        elif cl in ("price", "quantity"):
            v = _round_decimal(_to_float(v), digits=6)
        else:
            # Общее поведение для дат/даттаймов
            if hasattr(v, "isoformat") and callable(getattr(v, "isoformat")):
                try:
                    v = v.isoformat(sep=" ")
                except TypeError:
                    v = str(v)
        out.append(None if (v != 0 and not v) else v)
    return tuple(out)


def insert_rows(conn: sqlite3.Connection, table: str, columns: Sequence[str], rows: List[Dict[str, object]]) -> int:
    if not rows:
        return 0
    table_q = _quote_ident(_norm_table_name(table))
    cols_q = ", ".join(_quote_ident(c) for c in columns)
    placeholders = ", ".join(["?"] * len(columns))
    # OR IGNORE для пропуска дубликатов по уникальному индексу
    sql = f"INSERT OR IGNORE INTO {table_q} ({cols_q}) VALUES ({placeholders})"
    data = [_normalize_row_values(r, columns) for r in rows]
    # Подсчёт фактически вставленных строк через разницу COUNT(*)
    cur = conn.cursor()
    cur.execute(f"SELECT COUNT(*) FROM {table_q}")
    before_cnt = cur.fetchone()[0]
    cur.executemany(sql, data)
    conn.commit()
    cur.execute(f"SELECT COUNT(*) FROM {table_q}")
    after_cnt = cur.fetchone()[0]
    return int(after_cnt - before_cnt)


def _norm_str(val: object) -> str:
    if val is None:
        return ""
    try:
        return str(val).strip()
    except Exception:
        return ""


def _find_col_case_insensitive(columns: Sequence[str], target: str) -> Optional[str]:
    tl = target.lower()
    for c in columns:
        if c.lower() == tl:
            return c
    return None


def _is_garbage_record(mapping: Dict[str, object]) -> bool:
    # Проверяем признак мусорной строки: Event=CUSTOM_HOLDING_SETTINGS и Symbol=VK-ГДР
    # по именам столбцов без учёта регистра
    lowered = {str(k).lower(): v for k, v in mapping.items()}
    ev = _norm_str(lowered.get("event"))
    sym = _norm_str(lowered.get("symbol"))
    return ev.upper() == "CUSTOM_HOLDING_SETTINGS" and sym == "VK-ГДР"


def _load_excel_file_core(file_path: str, sheet: Optional[Union[str, int]], schema_columns: List[str]) -> Tuple[List[Dict[str, object]], int]:
    pd = _try_import_pandas()
    if pd is not None:
        sheet_arg = 0 if sheet is None else sheet
        df = pd.read_excel(file_path, sheet_name=sheet_arg)
        for col in schema_columns:
            if col not in df.columns:
                df[col] = None
        extra_cols = [c for c in df.columns if c not in schema_columns]
        if extra_cols:
            df = df.drop(columns=extra_cols)
        df = df[schema_columns]
        # Фильтр мусорной строки (Event=CUSTOM_HOLDING_SETTINGS и Symbol=VK-ГДР)
        ev_col = _find_col_case_insensitive(df.columns, "Event")
        sym_col = _find_col_case_insensitive(df.columns, "Symbol")
        skipped = 0
        if ev_col and sym_col:
            before = len(df)
            df = df[~(
                df[ev_col].astype(str).str.strip().str.upper().eq("CUSTOM_HOLDING_SETTINGS") &
                df[sym_col].astype(str).str.strip().eq("VK-ГДР")
            )]
            skipped = before - len(df)
        for col in df.columns:
            if str(df[col].dtype).startswith("datetime64"):
                df[col] = df[col].dt.strftime("%Y-%m-%d %H:%M:%S")
        df = df.where(pd.notnull(df), None)
        records = [dict(zip(schema_columns, row)) for row in df.itertuples(index=False, name=None)]
        return records, skipped

    ox = _try_import_openpyxl()
    if ox is None:
        raise RuntimeError("Не удалось импортировать pandas или openpyxl. Установите пакет: pip install pandas openpyxl")
    wb = ox.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet] if isinstance(sheet, str) else (wb.worksheets[sheet] if isinstance(sheet, int) else wb.active)
    rows: List[Dict[str, object]] = []
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(c) if c is not None else f"col_{i+1}" for i, c in enumerate(header)]
    skipped = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        rec = {h: v for h, v in zip(headers, r)}
        aligned = {c: rec.get(c) for c in schema_columns}
        if _is_garbage_record(aligned):
            skipped += 1
            continue
        rows.append(aligned)
    wb.close()
    return rows, skipped


def _stat_signature(path: str) -> Tuple[int, int]:
    st = os.stat(path)
    return int(st.st_size), int(getattr(st, "st_mtime_ns", int(st.st_mtime * 1e9)))


def load_excel_file(
    file_path: str,
    sheet: Optional[Union[str, int]],
    schema_columns: List[str],
    *,
    immutable_check: bool = True,
    safe_copy: bool = False,
) -> Tuple[List[Dict[str, object]], int]:
    """
    Загружает данные из Excel.
    - immutable_check: проверяет, что исходный файл не изменился по размеру/mtime во время чтения.
    - safe_copy: читает из временной копии файла.
    """
    original_path = file_path
    before_sig = _stat_signature(original_path)

    read_path = original_path
    tmp_path: Optional[str] = None
    try:
        if safe_copy:
            tmp_dir = tempfile.mkdtemp(prefix="xlsx_stage_")
            tmp_path = os.path.join(tmp_dir, os.path.basename(original_path))
            shutil.copy2(original_path, tmp_path)
            read_path = tmp_path

        records, skipped = _load_excel_file_core(read_path, sheet, schema_columns)
        return records, skipped
    finally:
        # Проверка неизменности исходного файла
        try:
            after_sig = _stat_signature(original_path)
            if immutable_check and after_sig != before_sig:
                raise SystemExit(
                    f"Обнаружено изменение исходного файла во время чтения (size/mtime). Файл: {original_path}"
                )
            elif after_sig != before_sig:
                print(
                    f"Внимание: исходный файл изменился во время чтения (size/mtime): {original_path}"
                )
        except FileNotFoundError:
            if immutable_check:
                raise SystemExit(f"Исходный файл исчез во время чтения: {original_path}")
            else:
                print(f"Внимание: исходный файл исчез во время чтения: {original_path}")
        finally:
            if tmp_path:
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass
                try:
                    os.rmdir(os.path.dirname(tmp_path))
                except OSError:
                    pass


def discover_xlsx_files(directory: str, exclude=()) -> List[str]:
    files: List[str] = []
    for name in os.listdir(directory):
        if name.lower().endswith(".xlsx") and not name.startswith("~$") and name not in exclude:
            files.append(os.path.join(directory, name))
    return sorted(files)


def main():
    default_dir = os.path.dirname(os.path.abspath(__file__))
    default_schema = os.path.join(default_dir, "STANDART_22_10.xlsx")
    default_db = os.path.join(default_dir, "operations.db")
    default_input_dir = os.path.join(default_dir, "input_xlsx")

    parser = argparse.ArgumentParser(
        description=(
            "Создаёт SQLite БД по схеме из стандартного .xlsx и загружает в неё операции из .xlsx файлов."
        )
    )
    parser.add_argument("--schema", dest="schema", default=default_schema, help="Путь к файлу-схеме .xlsx")
    parser.add_argument("--db", dest="db", default=default_db, help="Путь к файлу SQLite БД")
    parser.add_argument("--table", dest="table", default="operations", help="Имя таблицы в БД")
    parser.add_argument("--sheet", dest="sheet", default=None, help="Имя листа Excel (по умолчанию первый лист)")
    parser.add_argument(
        "--input-dir",
        dest="input_dir",
        default=default_input_dir,
        help="Папка с .xlsx файлами для загрузки (по умолчанию input_xlsx рядом со скриптом)",
    )
    parser.add_argument(
        "--files",
        nargs="*",
        default=None,
        help="Список .xlsx файлов для загрузки. Если не задан, загружаются все .xlsx из --input-dir (кроме schema)",
    )
    parser.add_argument(
        "--if-exists",
        choices=["append", "replace", "fail"],
        default="append",
        help="Поведение при существующей таблице: append/replace/fail (по умолчанию append)",
    )
    parser.add_argument(
        "--allow-replace",
        dest="allow_replace",
        action="store_true",
        help="Явно разрешить пересоздание таблицы при --if-exists replace",
    )
    parser.add_argument(
        "--safe-copy",
        dest="safe_copy",
        action="store_true",
        help="Читать входные .xlsx из временной копии для полной безопасности.",
    )
    parser.add_argument(
        "--no-immutable-check",
        dest="immutable_check",
        action="store_false",
        help="Отключить проверку неизменности исходных файлов (не рекомендуется)",
    )
    parser.set_defaults(immutable_check=True)
    parser.add_argument(
        "--allow-schema-as-data",
        dest="allow_schema_as_data",
        action="store_true",
        help="Разрешить использовать файл-схему как входной файл данных",
    )
    parser.add_argument(
        "--unique-on",
        dest="unique_on",
        default="Event,Date,Symbol,Quantity,Price",
        help=(
            "Столбцы для уникального индекса (через запятую). "
            "По умолчанию: Event,Date,Symbol,Quantity,Price. Также поддерживается 'all' или 'none'."
        ),
    )

    args = parser.parse_args()

    schema_path = os.path.abspath(args.schema)
    db_path = os.path.abspath(args.db)
    table = args.table
    sheet = args.sheet
    input_dir = os.path.abspath(args.input_dir)

    if not os.path.isfile(schema_path):
        raise SystemExit(f"Файл схемы не найден: {schema_path}")

    if not os.path.isdir(input_dir):
        os.makedirs(input_dir, exist_ok=True)

    schema_columns, schema_types = read_schema_from_excel(schema_path, sheet)

    # Разбор параметра уникальности
    unique_cols: Optional[List[str]]
    unique_on_raw = (args.unique_on or "").strip().lower()
    if unique_on_raw in ("", "none"):
        unique_cols = None
    elif unique_on_raw == "all":
        unique_cols = list(schema_columns)
    else:
        # список столбцов через запятую, сохраняем оригинальные имена (регистр"); проверяем существование
        requested = [c.strip() for c in args.unique_on.split(",") if c.strip()]
        not_found = [c for c in requested if c not in schema_columns]
        if not_found:
            raise SystemExit(
                "В --unique-on указаны столбцы, отсутствующие в схеме: " + ", ".join(not_found)
            )
        unique_cols = requested

    if args.files:
        input_files = [os.path.abspath(p) for p in args.files]
    else:
        exclude = [os.path.basename(schema_path)]
        input_files = discover_xlsx_files(input_dir, exclude=exclude)

    # Исключаем файл-схему из загрузки, если не разрешено явно
    if not args.allow_schema_as_data:
        input_files = [p for p in input_files if os.path.abspath(p) != os.path.abspath(schema_path)]

    if not input_files:
        print("Не найдены .xlsx файлы для загрузки в указанной папке.")
        print(f"Папка: {input_dir}")
        return

    conn = sqlite3.connect(db_path)
    try:
        if args.if_exists == "fail":
            cur = conn.cursor()
            cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name = ?", (_norm_table_name(table),))
            if cur.fetchone():
                raise SystemExit(f"Таблица '{table}' уже существует. Укажите --if-exists append/replace или другое имя.")

        ensure_table(conn, table, schema_columns, schema_types, args.if_exists)
        ensure_unique_index(conn, table, unique_cols)

        total_inserted = 0
        for fp in input_files:
            try:
                rows, skipped = load_excel_file(
                    fp,
                    sheet,
                    schema_columns,
                    immutable_check=args.immutable_check,
                    safe_copy=args.safe_copy,
                )
                inserted = insert_rows(conn, table, schema_columns, rows)
                total_inserted += inserted
                duplicates = max(0, len(rows) - inserted)
                print(
                    f"Файл: {fp}\n  добавлено: {inserted}\n  дубликатов: {duplicates}\n  отфильтровано (мусор): {skipped}"
                )
            except Exception as e:
                print(f"Ошибка загрузки файла {fp}: {e}")

        print("Готово.")
        print(f"БД: {db_path}")
        print(f"Таблица: {table}")
        print(f"Столбцы: {', '.join(schema_columns)}")
        print(f"Всего загружено строк: {total_inserted}")
    finally:
        conn.close()


def main2():
    default_dir = os.path.dirname(os.path.abspath(__file__))
    default_schema = os.path.join(default_dir, "STANDART_22_10.xlsx")
    default_db = os.path.join(default_dir, "operations.db")
    default_input_dir = os.path.join(default_dir, "input_xlsx")

    parser = argparse.ArgumentParser(
        description=(
            "Создаёт SQLite БД по схеме из стандартного .xlsx и загружает в неё операции из .xlsx файлов."
        )
    )
    parser.add_argument("--schema", dest="schema", default=default_schema, help="Путь к файлу-схеме .xlsx")
    parser.add_argument("--db", dest="db", default=default_db, help="Путь к файлу SQLite БД")
    parser.add_argument("--table", dest="table", default="operations", help="Имя таблицы в БД")
    parser.add_argument("--sheet", dest="sheet", default=None, help="Имя листа Excel (по умолчанию первый лист)")
    parser.add_argument(
        "--input-dir",
        dest="input_dir",
        default=default_input_dir,
        help="Папка с .xlsx файлами для загрузки (по умолчанию input_xlsx рядом со скриптом)",
    )
    parser.add_argument(
        "--files",
        nargs="*",
        default=None,
        help="Список .xlsx файлов для загрузки. Если не задан, загружаются все .xlsx из --input-dir (кроме schema)",
    )
    parser.add_argument(
        "--if-exists",
        choices=["append", "replace", "fail"],
        default="append",
        help="Поведение при существующей таблице: append/replace/fail (по умолчанию append)",
    )
    parser.add_argument(
        "--allow-replace",
        dest="allow_replace",
        action="store_true",
        help="Явно разрешить пересоздание таблицы при --if-exists replace",
    )
    parser.add_argument(
        "--safe-copy",
        dest="safe_copy",
        action="store_true",
        help="Читать входные .xlsx из временной копии для полной безопасности.",
    )
    parser.add_argument(
        "--no-immutable-check",
        dest="immutable_check",
        action="store_false",
        help="Отключить проверку неизменности исходных файлов (не рекомендуется)",
    )
    parser.set_defaults(immutable_check=True)
    parser.add_argument(
        "--allow-schema-as-data",
        dest="allow_schema_as_data",
        action="store_true",
        help="Разрешить использовать файл-схему как входной файл данных",
    )
    parser.add_argument(
        "--unique-on",
        dest="unique_on",
        default="all",
        help=(
            "Столбцы для уникального индекса (через запятую). "
            "Значения: 'all' (все столбцы), 'none' (без индекса), или список имён столбцов."
        ),
    )

    args = parser.parse_args()

    schema_path = os.path.abspath(args.schema)
    db_path = os.path.abspath(args.db)
    table = args.table
    sheet = args.sheet
    input_dir = os.path.abspath(args.input_dir)

    if not os.path.isfile(schema_path):
        raise SystemExit(f"Файл схемы не найден: {schema_path}")

    if not os.path.isdir(input_dir):
        os.makedirs(input_dir, exist_ok=True)

    schema_columns, schema_types = read_schema_from_excel(schema_path, sheet)

    if args.files:
        input_files = [os.path.abspath(p) for p in args.files]
    else:
        exclude = [os.path.basename(schema_path)]
        input_files = discover_xlsx_files(input_dir, exclude=exclude)

    if not args.allow_schema_as_data:
        input_files = [p for p in input_files if os.path.abspath(p) != os.path.abspath(schema_path)]

    if not input_files:
        print("Не найдены .xlsx файлы для загрузки в указанной папке.")
        print(f"Папка: {input_dir}")
        return

    conn = sqlite3.connect(db_path)
    try:
        tbl_exists = table_exists(conn, table)

        if args.if_exists == "fail" and tbl_exists:
            raise SystemExit(
                f"Таблица '{table}' уже существует. Укажите --if-exists append/replace или другое имя."
            )

        if tbl_exists:
            if args.if_exists == "replace":
                if not args.allow_replace:
                    raise SystemExit(
                        "Пересоздание таблицы запрещено по умолчанию. Укажите --allow-replace, если хотите заменить таблицу."
                    )
                ensure_table(conn, table, schema_columns, schema_types, args.if_exists)
                actual_columns, actual_types = list(schema_columns), dict(schema_types)
            else:
                actual_columns, actual_types = get_db_table_schema(conn, table)
        else:
            ensure_table(conn, table, schema_columns, schema_types, args.if_exists)
            actual_columns, actual_types = list(schema_columns), dict(schema_types)

        unique_on_raw = (args.unique_on or "").strip().lower()
        if unique_on_raw in ("", "none"):
            unique_final = None
        elif unique_on_raw == "all":
            unique_final = list(actual_columns)
        else:
            requested = [c.strip() for c in args.unique_on.split(",") if c.strip()]
            not_found = [c for c in requested if c not in actual_columns]
            if not_found:
                raise SystemExit(
                    "В --unique-on указаны столбцы, отсутствующие в таблице: " + ", ".join(not_found)
                )
            unique_final = requested

        ensure_unique_index(conn, table, unique_final)

        total_inserted = 0
        for fp in input_files:
            try:
                rows, skipped = load_excel_file(
                    fp,
                    sheet,
                    actual_columns,
                    immutable_check=args.immutable_check,
                    safe_copy=args.safe_copy,
                )
                inserted = insert_rows(conn, table, actual_columns, rows)
                total_inserted += inserted
                duplicates = max(0, len(rows) - inserted)
                print(
                    f"Файл: {fp}\n  добавлено: {inserted}\n  дубликатов: {duplicates}\n  отфильтровано (мусор): {skipped}"
                )
            except Exception as e:
                print(f"Ошибка загрузки файла {fp}: {e}")

        print("Готово.")
        print(f"БД: {db_path}")
        print(f"Таблица: {table}")
        print(f"Столбцы: {', '.join(actual_columns)}")
        print(f"Всего загружено строк: {total_inserted}")
    finally:
        conn.close()


# Переопределяем точку входа на версию с фиксированной схемой
main = main2

if __name__ == "__main__":
    main()
